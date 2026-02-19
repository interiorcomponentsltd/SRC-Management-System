#!/usr/bin/env python3
"""Better extraction handling for attendance data with formulas."""

import openpyxl
import json

def inspect_attendance_file():
    """Inspect the attendance file structure in detail."""
    try:
        wb = openpyxl.load_workbook('SPIRIT REALM CENTRE SUNDAYS ATTENDANCE TRACKER.xlsx')
        print(f"Sheet names: {wb.sheetnames}\n")
        
        # Check first sheet
        ws = wb.active
        print(f"Active sheet: {ws.title}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}\n")
        
        # Check for data_only version (to get values instead of formulas)
        print("Attempting to load with data_only=True...")
        wb_values = openpyxl.load_workbook('SPIRIT REALM CENTRE SUNDAYS ATTENDANCE TRACKER.xlsx', data_only=True)
        ws_values = wb_values.active
        
        # Get sample data with values
        print("\nFirst 10 rows with values_only=True:")
        for idx, row in enumerate(ws_values.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True), 1):
            print(f"Row {idx}: {row}")
        
        # Count non-empty names
        names_found = 0
        for row in ws_values.iter_rows(min_row=3, values_only=True):
            if row[1] and str(row[1]).strip() and not str(row[1]).startswith('='):
                names_found += 1
        
        print(f"\nEstimated members with actual values: {names_found}")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect_attendance_file()
