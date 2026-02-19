#!/usr/bin/env python3
import openpyxl
import os

os.chdir(r'c:\Users\barth\OneDrive\Desktop\SRC ATTENDANCE ADMIN\src-mgt-sys')

print("=== FIRST TIMERS FILE ===")
wb1 = openpyxl.load_workbook('FIRST TIMERS NEW MEMBERS1.xlsx')
for sheet_name in wb1.sheetnames:
    print(f"\nSheet: {sheet_name}")
    ws = wb1[sheet_name]
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    # Print first 10 rows
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 10:
            print(f"Row {i}: {row}")
        else:
            break

print("\n" + "="*80)
print("=== ATTENDANCE TRACKER FILE ===")
wb2 = openpyxl.load_workbook('SPIRIT REALM CENTRE SUNDAYS ATTENDANCE TRACKER.xlsx')
for sheet_name in wb2.sheetnames:
    print(f"\nSheet: {sheet_name}")
    ws = wb2[sheet_name]
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    # Print first 10 rows
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 10:
            print(f"Row {i}: {row}")
        else:
            break
