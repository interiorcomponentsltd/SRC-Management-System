#!/usr/bin/env python3
"""Extract members from both Excel files with proper department assignment."""

import openpyxl
import json
import re

def extract_first_timers():
    """Extract first timers data."""
    try:
        wb = openpyxl.load_workbook('FIRST TIMERS NEW MEMBERS1.xlsx')
        ws = wb.active
        
        members = []
        member_id = 1
        
        # Data starts at row 3 (row 1: title, row 2: another title, row 3: actual headers with data)
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 1):
            if row[1] is None or str(row[1]).strip() == '':
                continue
            
            name = str(row[1]).strip()
            
            # Extract and clean phone number
            phone = str(row[2]).strip() if row[2] else ''
            # Remove non-digit characters except + and spaces
            phone = re.sub(r'[^\d\s+]', '', phone).strip()
            
            member = {
                'id': member_id,
                's_n': row[0] if row[0] else member_id,
                'name': name,
                'phone': phone,
                'marital_status': str(row[3]).strip() if row[3] else '',
                'gender': str(row[4]).strip().upper() if row[4] else '',
                'address': str(row[5]).strip() if row[5] else '',
                'state_of_origin': str(row[6]).strip() if row[6] else '',
                'occupation': str(row[7]).strip() if row[7] else '',
                'prayer_request': str(row[8]).strip() if row[8] else '',
                'decision': str(row[9]).strip() if row[9] else '',
                'invited_by': str(row[10]).strip() if row[10] else '',
                'counselor': str(row[11]).strip() if row[11] else '',
                'date_joined': str(row[12]).strip() if row[12] else '',
                'baptism': str(row[13]).strip() if row[13] else '',
                'status': 'first_timer',
                'department': '',
                'position': 'Member',
                'photo_url': ''
            }
            members.append(member)
            member_id += 1
        
        print(f"✓ Extracted {len(members)} first timer members")
        return members
    except Exception as e:
        print(f"✗ Error extracting first timers: {e}")
        return []

def extract_existing_members():
    """Extract existing members from department sheets."""
    try:
        wb = openpyxl.load_workbook('SPIRIT REALM CENTRE SUNDAYS ATTENDANCE TRACKER.xlsx', data_only=True)
        
        # Department sheet mapping (from the Excel file)
        dept_sheets = {
            'GENERAL': 'General Members',
            'MUSIC': 'Music House',
            'USHERS': 'Ushers',
            'MEDIA': 'Media',
            'SECURITY': 'Security',
            'PRAYER GROUP': 'Prayer Group',
            'GODLY ACTION': 'Godly Action',
            'COUNSELLING ': 'Counseling',
            'SANCTUARY': 'Sanctuary',
            'EWEKA': 'Men of Impact'  # Assuming EWEKA is Men of Impact group
        }
        
        all_members = []
        member_id = 1001  # Start from different range
        
        for sheet_name, dept_name in dept_sheets.items():
            try:
                if sheet_name not in wb.sheetnames:
                    print(f"  - Sheet '{sheet_name}' not found, skipping...")
                    continue
                    
                ws = wb[sheet_name]
                sheet_members = 0
                
                # Extract names from column B (index 1), starting from row 3
                for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=2, values_only=True), 1):
                    if row[1] is None:
                        continue
                    
                    name = str(row[1]).strip()
                    
                    # Skip empty, formula references, and header-like entries
                    if (not name or name.startswith('=') or 
                        name.upper() in ['NAME', 'NAMES', 'NAMES OF MEMBERS']):
                        continue
                    
                    member = {
                        'id': member_id,
                        's_n': row[0] if row[0] else member_id - 1000,
                        'name': name,
                        'phone': '',
                        'marital_status': '',
                        'gender': '',
                        'address': '',
                        'state_of_origin': '',
                        'occupation': '',
                        'prayer_request': '',
                        'decision': '',
                        'invited_by': '',
                        'counselor': '',
                        'date_joined': '',
                        'baptism': '',
                        'status': 'existing_member',
                        'department': dept_name,
                        'position': 'Member',
                        'photo_url': ''
                    }
                    all_members.append(member)
                    member_id += 1
                    sheet_members += 1
                
                print(f"  ✓ {dept_name}: {sheet_members} members")
                
            except Exception as e:
                print(f"  ✗ Error processing '{sheet_name}': {e}")
                continue
        
        return all_members
    except Exception as e:
        print(f"✗ Error extracting existing members: {e}")
        return []

def deduplicate_members(all_members):
    """Remove duplicate members based on name."""
    seen = {}
    unique = []
    
    for member in all_members:
        name_key = member['name'].lower().strip()
        if name_key not in seen:
            seen[name_key] = True
            unique.append(member)
    
    return unique

def assign_unassigned_departments(members):
    """Assign unassigned members to departments round-robin."""
    depts = [
        'Men of Impact',
        'Women of Purpose',
        'Music House',
        'Ushers',
        'Security',
        'Sanctuary',
        'Prayer Group',
        'Media',
        'Godly Action',
        'Counseling'
    ]
    
    unassigned_idx = 0
    for member in members:
        if not member.get('department') or member['department'] == '':
            member['department'] = depts[unassigned_idx % len(depts)]
            unassigned_idx += 1
    
    return members

# Main execution
if __name__ == "__main__":
    print("=" * 60)
    print("MEMBER DATA EXTRACTION")
    print("=" * 60)
    
    print("\nExtracting first timer members...")
    ft_members = extract_first_timers()
    
    print("\nExtracting existing members from departments...")
    ex_members = extract_existing_members()
    
    # Combine
    all_members = ft_members + ex_members
    print(f"\nTotal members before deduplication: {len(all_members)}")
    
    # Deduplicate
    unique_members = deduplicate_members(all_members)
    print(f"Total unique members: {len(unique_members)}")
    
    # Assign unassigned
    unique_members = assign_unassigned_departments(unique_members)
    
    # Reassign IDs
    for idx, member in enumerate(unique_members, 1):
        member['id'] = idx
    
    # Generate statistics
    print(f"\n" + "=" * 60)
    print("STATISTICS")
    print("=" * 60)
    
    first_timers = sum(1 for m in unique_members if m['status'] == 'first_timer')
    existing = sum(1 for m in unique_members if m['status'] == 'existing_member')
    
    print(f"Total members: {len(unique_members)}")
    print(f"First timers: {first_timers}")
    print(f"Existing members: {existing}")
    
    # Department distribution
    dept_dist = {}
    for member in unique_members:
        dept = member.get('department', 'Unassigned')
        dept_dist[dept] = dept_dist.get(dept, 0) + 1
    
    print(f"\nDepartment Distribution:")
    for dept, count in sorted(dept_dist.items()):
        print(f"  {dept}: {count} members")
    
    # Save to JSON
    with open('extracted_members.json', 'w') as f:
        json.dump(unique_members, f, indent=2, default=str)
    
    print(f"\n✓ Data saved to extracted_members.json")
    print("=" * 60)
