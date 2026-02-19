#!/usr/bin/env python3
"""
Extract and merge members from two Excel files:
1. FIRST TIMERS NEW MEMBERS1.xlsx (Sheet1, from row 4 onward, headers in row 3)
2. SPIRIT REALM CENTRE SUNDAYS ATTENDANCE TRACKER.xlsx (GENERAL sheet)

Generates:
- general_members.json: All members for general distribution
- first_timers_followup.json: First timers-only list for follow-ups
"""

import openpyxl
import json
import os
import re

def clean_name(name):
    """Clean and standardize names."""
    if not name:
        return ""
    name_str = str(name).strip()
    # Remove leading/trailing numbers and special chars
    name_str = re.sub(r'^\d+\s*', '', name_str)
    return name_str.upper()

def clean_phone(phone):
    """Clean phone numbers."""
    if not phone:
        return ""
    phone_str = str(phone).strip()
    # Remove non-digit chars
    phone_str = re.sub(r'\D', '', phone_str)
    return phone_str[-10:] if len(phone_str) >= 10 else phone_str

def extract_first_timers():
    """Extract first timers from Sheet1 (headers in row 3, data from row 4)"""
    print("Extracting first timers from Sheet1...")
    try:
        wb = openpyxl.load_workbook('FIRST TIMERS NEW MEMBERS1.xlsx')
        ws = wb['Sheet1']
        
        first_timers = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            s_n, name, phone, marital_status, gender, address, state, occupation, prayer_request, decision, invited_by, counselor, date, baptism = row[:14]
            
            name_clean = clean_name(name)
            if not name_clean:
                continue
            
            member = {
                "id": 10000 + row_idx,
                "name": name_clean,
                "phone": clean_phone(phone),
                "marital_status": str(marital_status).strip() if marital_status else "",
                "gender": str(gender).strip().upper() if gender else "",
                "address": str(address).strip() if address else "",
                "state_of_origin": str(state).strip() if state else "",
                "occupation": str(occupation).strip() if occupation else "",
                "prayer_request": str(prayer_request).strip() if prayer_request else "",
                "invited_by": str(invited_by).strip() if invited_by else "",
                "date_joined": str(date).strip() if date else "",
                "status": "first_timer",
                "department": "Unassigned",
                "source": "first_timers_excel"
            }
            first_timers.append(member)
        
        print(f"  Extracted {len(first_timers)} first timers")
        return first_timers
    except Exception as e:
        print(f"  Error: {e}")
        return []

def extract_general_sheet():
    """Extract members from GENERAL sheet in attendance tracker"""
    print("Extracting members from GENERAL sheet...")
    try:
        wb = openpyxl.load_workbook('SPIRIT REALM CENTRE SUNDAYS ATTENDANCE TRACKER.xlsx')
        ws = wb['GENERAL']
        
        members = {}  # Use dict to deduplicate by name
        member_id = 1000
        
        # Start from row 8 (row 8 onwards contain member data in column B)
        for row_idx in range(8, ws.max_row + 1):
            cell_b = ws[f'B{row_idx}'].value
            if not cell_b or isinstance(cell_b, str) and cell_b.startswith('='):
                continue
            
            # Parse member info: "NAME; PHONE; [Department]; [Address]"
            member_str = str(cell_b).strip()
            if not member_str or member_str.startswith('MEMBER'):
                continue
            
            # Split by semicolon
            parts = [p.strip() for p in member_str.split(';')]
            name = clean_name(parts[0]) if len(parts) > 0 else ""
            
            if not name or name in members:
                continue
            
            phone = clean_phone(parts[1]) if len(parts) > 1 else ""
            dept = parts[2].strip().replace('[', '').replace(']', '') if len(parts) > 2 else "Unassigned"
            address = parts[3].strip().replace('[', '').replace(']', '') if len(parts) > 3 else ""
            
            members[name] = {
                "id": member_id,
                "name": name,
                "phone": phone,
                "address": address,
                "gender": "",
                "status": "member",
                "department": dept if dept not in ['GO Headquarters', 'Headquarters'] else "Unassigned",
                "source": "attendance_excel"
            }
            member_id += 1
        
        members_list = list(members.values())
        print(f"  Extracted {len(members_list)} unique members")
        return members_list
    except Exception as e:
        print(f"  Error: {e}")
        return []

def dedup_by_name(list1, list2):
    """Merge two lists, deduplicating by name  (list1 takes precedence)"""
    names_seen = {m["name"] for m in list1}
    merged = list1.copy()
    for member in list2:
        if member["name"] not in names_seen:
            merged.append(member)
            names_seen.add(member["name"])
    return merged

def main():
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    
    # Extract data
    first_timers = extract_first_timers()
    general_members = extract_general_sheet()
    
    # Merge: general members first, then add non-duplicate first timers
    all_members = dedup_by_name(general_members, first_timers)
    
    # Assign sequential IDs
    for idx, member in enumerate(all_members, start=1):
        member["id"] = idx
    
    print(f"\nMerged: {len(all_members)} total unique members")
    
    # Save general members list
    general_members_file = "general_members.json"
    with open(general_members_file, "w", encoding="utf-8") as f:
        json.dump(all_members, f, indent=2, ensure_ascii=False)
    print(f"Saved to {general_members_file}")
    
    # Save first timers follow-up list
    first_timers_file = "first_timers_followup.json"
    with open(first_timers_file, "w", encoding="utf-8") as f:
        json.dump(first_timers, f, indent=2, ensure_ascii=False)
    print(f"Saved to {first_timers_file}")
    
    # Print summary
    print(f"\n=== Summary ===")
    print(f"Total Members: {len(all_members)}")
    print(f"First Timers: {len(first_timers)}")
    print(f"General Members: {len(general_members)}")
    
    # Show sample
    if all_members:
        print(f"\nSample member:\n{json.dumps(all_members[0], indent=2, ensure_ascii=False)}")

if __name__ == "__main__":
    main()
