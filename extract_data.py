#!/usr/bin/env python3
"""Extract data from Excel files and prepare for app integration."""

import openpyxl
import json
from datetime import datetime

def extract_first_timers():
    """Extract first timers data from Excel file."""
    try:
        wb = openpyxl.load_workbook('FIRST TIMERS NEW MEMBERS1.xlsx')
        ws = wb.active
        
        # Get header row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        print("=== FIRST TIMERS NEW MEMBERS ===")
        print(f"Columns found: {headers}")
        print(f"\nTotal columns: {len(headers)}")
        
        # Get first few data rows
        print("\nFirst 3 data rows:")
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), 1):
            print(f"Row {idx}: {row}")
        
        # Count total records
        total_rows = ws.max_row - 1  # Exclude header
        print(f"\nTotal records: {total_rows}")
        
        return headers, total_rows
    except Exception as e:
        print(f"Error extracting first timers: {e}")
        return None, None

def extract_attendance():
    """Extract attendance data from Excel file."""
    try:
        wb = openpyxl.load_workbook('SPIRIT REALM CENTRE SUNDAYS ATTENDANCE TRACKER.xlsx')
        ws = wb.active
        
        # Get header row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        print("\n\n=== SPIRIT REALM CENTRE SUNDAYS ATTENDANCE TRACKER ===")
        print(f"Columns found: {headers}")
        print(f"\nTotal columns: {len(headers)}")
        
        # Get first few data rows
        print("\nFirst 3 data rows:")
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), 1):
            print(f"Row {idx}: {row}")
        
        # Count total records
        total_rows = ws.max_row - 1  # Exclude header
        print(f"\nTotal records: {total_rows}")
        
        return headers, total_rows
    except Exception as e:
        print(f"Error extracting attendance: {e}")
        return None, None

if __name__ == "__main__":
    extract_first_timers()
    extract_attendance()
